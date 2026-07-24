"""Escrita de dados extraídos em planilhas .xlsx."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def write_rows(
    output_path: str | Path,
    fields: list[str],
    rows: list[dict],
    sheet_name: str = "Extração",
) -> Path:
    """Cria um .xlsx com `fields` como colunas e `rows` como linhas."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    for col, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, field in enumerate(fields, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(field, ""))

    # Ajuste simples de largura das colunas
    for col in range(1, len(fields) + 1):
        letter = get_column_letter(col)
        max_len = max(
            [len(str(fields[col - 1]))]
            + [len(str(row.get(fields[col - 1], ""))) for row in rows]
            or [10]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)

    wb.save(output_path)
    return output_path
